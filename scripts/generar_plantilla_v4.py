import sys
sys.path.insert(0, ".")

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from scripts.parse_mapa_practica import parse_mapa_practica

PLANTILLA_ORIGEN = "data/Plantilla_V3_FacSalud (1) (2).xlsx"
MAPA_PRACTICA = "data/info_reunion_refinacion_modelo/Mapa de practica general Medicina 2025-1.xlsx"
SALIDA = "data/Plantilla_V4_Refinada.xlsx"

print("Generando Plantilla V4 Refinada...")

wb_origen = load_workbook(PLANTILLA_ORIGEN)
wb_origen.save(SALIDA)
print(f"✓ Copiadas hojas: {wb_origen.sheetnames}")

rotaciones = parse_mapa_practica(MAPA_PRACTICA)
rotaciones = rotaciones[rotaciones["Semestre_plan"].between(5, 10)].copy()
rotaciones = rotaciones.rename(columns={
    "Semestre_plan": "Semestre_Plan",
    "ID_Institucion": "ID_Institucion",
    "Institucion": "Institucion",
    "Sede": "Sede",
    "Asignatura": "Asignatura",
    "Rotacion": "Rotacion",
    "Cupo": "Cupo_Maximo",
})
rotaciones = rotaciones[["Semestre_Plan", "Asignatura", "Rotacion", "ID_Institucion", "Institucion", "Sede", "Cupo_Maximo"]]

print(f"✓ Rotaciones 5to-10mo: {len(rotaciones)} registros")

demanda_data = [
    {"Semestre_Plan": 5, "Programa": "Medicina", "Tipo_Estudiante": "Pregrado", "Demanda_Estudiantes": 72, "Grupo_Min": 4, "Grupo_Max": 7, "Techo_Max": 75, "Observaciones": ""},
    {"Semestre_Plan": 6, "Programa": "Medicina", "Tipo_Estudiante": "Pregrado", "Demanda_Estudiantes": 73, "Grupo_Min": 4, "Grupo_Max": 7, "Techo_Max": 75, "Observaciones": ""},
    {"Semestre_Plan": 7, "Programa": "Medicina", "Tipo_Estudiante": "Pregrado", "Demanda_Estudiantes": 46, "Grupo_Min": 4, "Grupo_Max": 7, "Techo_Max": 75, "Observaciones": ""},
    {"Semestre_Plan": 8, "Programa": "Medicina", "Tipo_Estudiante": "Pregrado", "Demanda_Estudiantes": 56, "Grupo_Min": 4, "Grupo_Max": 7, "Techo_Max": 75, "Observaciones": ""},
    {"Semestre_Plan": 9, "Programa": "Medicina", "Tipo_Estudiante": "Pregrado", "Demanda_Estudiantes": 57, "Grupo_Min": 3, "Grupo_Max": 5, "Techo_Max": 75, "Observaciones": ""},
    {"Semestre_Plan": 10, "Programa": "Medicina", "Tipo_Estudiante": "Pregrado", "Demanda_Estudiantes": 36, "Grupo_Min": 3, "Grupo_Max": 5, "Techo_Max": 75, "Observaciones": ""},
]
demanda = pd.DataFrame(demanda_data)

print(f"✓ Demanda: {len(demanda)} semestres")

sets_ponderaciones = [
    {
        "set_id": "SET-SEM5-SaludPublica",
        "nombre": "Sem 5 - Salud Pública y Comunidad",
        "descripcion": "Prioriza IPS con servicios pediátricos, obstetricia, hospitales universitarios y escenarios avalados",
        "criterios": {
            "Es_Hospital_Universitario": 0.15,
            "Escenario_Avalado_Practicas": 0.20,
            "Servicios_Pediatricos (0/1)": 0.20,
            "Servicios_Obstetricia (0/1)": 0.15,
            "MisionVisionProposito_AlineacionDocencia (1-5)": 0.05,
            "Admiten_Docentes_Externos (Sí/No)": 0.05,
            "Areas_Bienestar (0/1)": 0.05,
            "Areas_Academicas (0/1)": 0.05,
            "%_Contraprestacion_Matricula (0-100)": 0.05,
            "EPP_Exigidos (Sin exigencia/Parcial/Completo + detalle)": 0.05,
        }
    },
    {
        "set_id": "SET-SEM6-Psiquiatria",
        "nombre": "Sem 6 - Psiquiatría y Enlace",
        "descripcion": "Prioriza IPS con evaluación, docentes externos, áreas académicas",
        "criterios": {
            "Es_Hospital_Universitario": 0.20,
            "Escenario_Avalado_Practicas": 0.20,
            "Servicios_Pediatricos (0/1)": 0.02,
            "Servicios_Obstetricia (0/1)": 0.03,
            "MisionVisionProposito_AlineacionDocencia (1-5)": 0.15,
            "Admiten_Docentes_Externos (Sí/No)": 0.15,
            "Areas_Bienestar (0/1)": 0.10,
            "Areas_Academicas (0/1)": 0.10,
            "%_Contraprestacion_Matricula (0-100)": 0.03,
            "EPP_Exigidos (Sin exigencia/Parcial/Completo + detalle)": 0.02,
        }
    },
    {
        "set_id": "SET-SEM7-MedicinaInterna",
        "nombre": "Sem 7 - Medicina Interna y Especialidades",
        "descripcion": "Prioriza IPS con vinculación de especialistas, evaluación, académicos",
        "criterios": {
            "Es_Hospital_Universitario": 0.20,
            "Escenario_Avalado_Practicas": 0.15,
            "Servicios_Pediatricos (0/1)": 0.05,
            "Servicios_Obstetricia (0/1)": 0.05,
            "MisionVisionProposito_AlineacionDocencia (1-5)": 0.10,
            "Admiten_Docentes_Externos (Sí/No)": 0.10,
            "Areas_Bienestar (0/1)": 0.10,
            "Areas_Academicas (0/1)": 0.15,
            "%_Contraprestacion_Matricula (0-100)": 0.05,
            "EPP_Exigidos (Sin exigencia/Parcial/Completo + detalle)": 0.05,
        }
    },
    {
        "set_id": "SET-SEM8-Pediatria",
        "nombre": "Sem 8 - Pediatría",
        "descripcion": "Prioriza IPS con servicios pediátricos, obstetricia, bienestar",
        "criterios": {
            "Es_Hospital_Universitario": 0.15,
            "Escenario_Avalado_Practicas": 0.15,
            "Servicios_Pediatricos (0/1)": 0.25,
            "Servicios_Obstetricia (0/1)": 0.15,
            "MisionVisionProposito_AlineacionDocencia (1-5)": 0.05,
            "Admiten_Docentes_Externos (Sí/No)": 0.05,
            "Areas_Bienestar (0/1)": 0.10,
            "Areas_Academicas (0/1)": 0.05,
            "%_Contraprestacion_Matricula (0-100)": 0.03,
            "EPP_Exigidos (Sin exigencia/Parcial/Completo + detalle)": 0.02,
        }
    },
    {
        "set_id": "SET-SEM9-Gineco",
        "nombre": "Sem 9 - Ginecobstetricia y Urgencias",
        "descripcion": "Prioriza IPS con obstetricia, quirófanos, urgencias, bajo costo",
        "criterios": {
            "Es_Hospital_Universitario": 0.15,
            "Escenario_Avalado_Practicas": 0.15,
            "Servicios_Pediatricos (0/1)": 0.10,
            "Servicios_Obstetricia (0/1)": 0.25,
            "MisionVisionProposito_AlineacionDocencia (1-5)": 0.05,
            "Admiten_Docentes_Externos (Sí/No)": 0.05,
            "Areas_Bienestar (0/1)": 0.05,
            "Areas_Academicas (0/1)": 0.05,
            "%_Contraprestacion_Matricula (0-100)": 0.10,
            "EPP_Exigidos (Sin exigencia/Parcial/Completo + detalle)": 0.05,
        }
    },
    {
        "set_id": "SET-SEM10-Cirugia",
        "nombre": "Sem 10 - Cirugía y Urgencias",
        "descripcion": "Prioriza IPS con quirófanos, especialistas, menor costo, medicina de urgencias",
        "criterios": {
            "Es_Hospital_Universitario": 0.20,
            "Escenario_Avalado_Practicas": 0.15,
            "Servicios_Pediatricos (0/1)": 0.05,
            "Servicios_Obstetricia (0/1)": 0.10,
            "MisionVisionProposito_AlineacionDocencia (1-5)": 0.05,
            "Admiten_Docentes_Externos (Sí/No)": 0.10,
            "Areas_Bienestar (0/1)": 0.05,
            "Areas_Academicas (0/1)": 0.10,
            "%_Contraprestacion_Matricula (0-100)": 0.15,
            "EPP_Exigidos (Sin exigencia/Parcial/Completo + detalle)": 0.05,
        }
    },
]

rows_pond = []
for s in sets_ponderaciones:
    for criterio, peso in s["criterios"].items():
        if "Contraprestacion" in criterio or "EPP_Exigidos" in criterio:
            tipo = "Costo"
        else:
            tipo = "Beneficio"
        rows_pond.append({
            "Set_ID": s["set_id"],
            "Nombre_Set": s["nombre"],
            "Semestre_Vigencia (AAAA-S)": "2026-1",
            "Programa (o GLOBAL)": "Medicina",
            "Tipo_Estudiante (opcional)": "Pregrado",
            "Tipo_Practica (opcional)": "",
            "Criterio_Codigo": criterio,
            "Tipo (Beneficio/Costo)": tipo,
            "Peso (0-1)": peso,
            "Activo (0/1)": 1.0,
            "Notas": s["descripcion"],
            "Suma_Pesos_Activos_del_Set (auto)": "",
        })

ponderaciones = pd.DataFrame(rows_pond)
print(f"✓ Ponderaciones: {len(ponderaciones)} criterios en {ponderaciones['Set_ID'].nunique()} sets")

with pd.ExcelWriter(SALIDA, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    rotaciones.to_excel(writer, sheet_name="06_Rotaciones", index=False)
    demanda.to_excel(writer, sheet_name="07_Demanda_Semestres", index=False)
    ponderaciones.to_excel(writer, sheet_name="05_Ponderaciones", index=False, startrow=4)

print(f"✓ Hojas actualizadas: 05_Ponderaciones (6 sets), 06_Rotaciones, 07_Demanda_Semestres")

wb = load_workbook(SALIDA)

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

for sheet_name in ["06_Rotaciones", "07_Demanda_Semestres"]:
    ws = wb[sheet_name]
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 35)

    ws.freeze_panes = "A2"

ws_pond = wb["05_Ponderaciones"]
for cell in ws_pond[5]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col in ws_pond.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws_pond.column_dimensions[col_letter].width = min(max_length + 2, 35)

ws_pond.freeze_panes = "A6"

wb.save(SALIDA)
print(f"\n✅ Archivo generado: {SALIDA}")

print("\n=== ESTRUCTURA DEL ARCHIVO ===")
for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f"  {sheet}: {ws.max_row} filas × {ws.max_column} columnas")

print("\n=== SETS DE PONDERACIONES ===")
for s in sets_ponderaciones:
    print(f"  {s['set_id']}: {s['nombre']}")
    total = sum(s['criterios'].values())
    print(f"    Suma pesos: {total:.4f}")

print("\n=== HOJA 07_Demanda_Semestres ===")
print(demanda.to_string(index=False))
