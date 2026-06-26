"""
Aplicación principal Streamlit
"""

import streamlit as st
import pandas as pd
import numpy as np
from pathlib import Path
import tempfile
from io import BytesIO
import logging
from typing import Optional, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Imports locales
from src.core import DataLoader, Optimizer, ScoreCalculator
from src.core.optimizer import GroupOptimizer
from src.utils import setup_logging
from src.visualization import (
    render_header, render_upload_section, render_config_section,
    render_results_summary, render_asignaciones_table, render_capacidad_chart,
    render_demanda_vs_asignacion, render_debug_info
)
from scripts.parse_mapa_practica import get_group_constraints

# Configurar logging
logger = setup_logging("logs", "debug_logs")

# Configurar estado de sesión
if "results" not in st.session_state:
    st.session_state.results = None
if "excel_data" not in st.session_state:
    st.session_state.excel_data = None


def generate_ejemplo_demanda(
    semestre: str = "2026-1",
    total_estudiantes: int = 80,
    programa: str = "Medicina",
    tipo_estudiante: str = "Pregrado",
    tipo_practica: str = "Rotación pregrado"
) -> pd.DataFrame:
    """Genera demanda de ejemplo (grupo único)"""
    return pd.DataFrame([
        {
            "Semestre": semestre,
            "Programa": programa,
            "Tipo_Estudiante": tipo_estudiante,
            "Tipo_Practica": tipo_practica,
            "Demanda_Estudiantes": int(total_estudiantes),
        }
    ])


def get_config_options_from_upload(uploaded_file):
    """Extrae Set_ID y semestres disponibles desde el archivo subido."""
    if uploaded_file is None:
        return [], []
    try:
        content = uploaded_file.getvalue()
        pond = pd.read_excel(BytesIO(content), sheet_name="05_Ponderaciones", header=4)
        set_options = (
            pond["Set_ID"].dropna().astype(str).str.strip().replace("", np.nan).dropna().unique().tolist()
            if "Set_ID" in pond.columns else []
        )
        semestre_options = (
            pond["Semestre_Vigencia (AAAA-S)"].dropna().astype(str).str.strip().replace("", np.nan).dropna().unique().tolist()
            if "Semestre_Vigencia (AAAA-S)" in pond.columns else []
        )
        return sorted(set_options), sorted(semestre_options)
    except Exception:
        return [], []


def get_asignaturas_for_semestre(uploaded_file, semestre_plan: int) -> list:
    """Lee 06_Rotaciones desde el archivo subido y retorna asignaturas del semestre."""
    if uploaded_file is None:
        return []
    try:
        content = uploaded_file.getvalue()
        rot = pd.read_excel(BytesIO(content), sheet_name="06_Rotaciones")
        col_sem = "Semestre_Plan" if "Semestre_Plan" in rot.columns else rot.columns[0]
        rot[col_sem] = pd.to_numeric(rot[col_sem], errors="coerce")
        rot = rot[rot[col_sem] == semestre_plan]
        if "Asignatura" not in rot.columns:
            return []
        asigs = (
            rot["Asignatura"]
            .dropna()
            .astype(str)
            .str.strip()
            .replace("", np.nan)
            .dropna()
            .unique()
            .tolist()
        )
        return sorted(asigs)
    except Exception:
        return []


def get_demanda_for_semestre(uploaded_file, semestre_plan: int) -> Optional[int]:
    """Lee 07_Demanda_Semestres y retorna Demanda_Estudiantes para el semestre dado."""
    if uploaded_file is None:
        return None
    try:
        content = uploaded_file.getvalue()
        df = pd.read_excel(BytesIO(content), sheet_name="07_Demanda_Semestres")
        if "Semestre_Plan" not in df.columns or "Demanda_Estudiantes" not in df.columns:
            return None
        df["Semestre_Plan"] = pd.to_numeric(df["Semestre_Plan"], errors="coerce")
        row = df[df["Semestre_Plan"] == semestre_plan]
        if row.empty:
            return None
        val = row["Demanda_Estudiantes"].iloc[0]
        return int(val) if pd.notna(val) else None
    except Exception:
        return None


def preview_capacidad(uploaded_file) -> int:
    """Calcula capacidad total estimada desde 02_Oferta_x_Programa."""
    if uploaded_file is None:
        return 0
    try:
        content = uploaded_file.getvalue()
        cupos = pd.read_excel(BytesIO(content), sheet_name="02_Oferta_x_Programa")
        if "Cupo_Estimado_Semestral" not in cupos.columns:
            return 0
        vals = pd.to_numeric(cupos["Cupo_Estimado_Semestral"], errors="coerce").fillna(0)
        return int(vals.sum())
    except Exception:
        return 0


def generate_ejemplo_cupos(semestre: str = "2026-1") -> pd.DataFrame:
    """Genera cupos de ejemplo"""
    inst_ids = [7600103715, 500102104, 7600102541, 7600103359, 7600108077]
    rows = []
    for inst in inst_ids:
        rows.append({
            "ID_Institucion": inst,
            "Programa": "Medicina",
            "Tipo_Estudiante (Pregrado/Posgrado)": "Pregrado",
            "Semestre (AAAA-S)": semestre,
            "Cupo_Estimado_Semestral": 15
        })
    return pd.DataFrame(rows)


def generate_ejemplo_costos(semestre: str = "2026-1") -> pd.DataFrame:
    """Genera costos de ejemplo"""
    inst_ids = [7600103715, 500102104, 7600102541, 7600103359, 7600108077]
    rows = []
    for inst in inst_ids:
        for tipo_pract in ["Rotación pregrado", "Internado de medicina"]:
            rows.append({
                "ID_Institucion": inst,
                "Programa_Costo": "Medicina",
                "Tipo_Estudiante_Costo": "Pregrado",
                "Tipo_Practica_Costo": tipo_pract,
                "Semestre_Vigencia (AAAA-S)": semestre,
                "%_Contraprestacion_Matricula (0-100)": 30.0,
                "Cobro_EPP (No cobra/Cobra a la Universidad)": "No cobra EPP",
            })
    return pd.DataFrame(rows)


def clean_criterio_codigo(code: str) -> str:
    """Limpia código de criterio removiendo sufijos como '(...)' y espacios extra."""
    text = str(code or "").strip()
    if " (" in text:
        text = text.split(" (", 1)[0]
    return " ".join(text.split())


def to_bool01(x) -> int:
    """Mapea valores comunes Sí/No a 1/0 de forma robusta."""
    if pd.isna(x):
        return 0
    if isinstance(x, (int, float, np.integer, np.floating)):
        return 1 if float(x) > 0 else 0
    value = str(x).strip().lower()
    if value in {"sí", "si", "1", "true", "yes"}:
        return 1
    if value in {"no", "0", "false"}:
        return 0
    return 0


def map_epp_exigidos(x):
    """Mapea EPP_Exigidos a costo numérico: 0 (sin), 0.5 (parcial), 1 (completo)."""
    if pd.isna(x):
        return np.nan
    value = str(x).strip().lower()
    if "sin exigencia" in value:
        return 0.0
    if "parcial" in value:
        return 0.5
    if "completo" in value:
        return 1.0
    return np.nan


def compute_scores_debug(score_rows: list) -> pd.DataFrame:
    """Devuelve score por institución y criterios sk para auditoría."""
    if not score_rows:
        return pd.DataFrame()
    df = pd.DataFrame(score_rows)
    value_cols = [c for c in df.columns if c.startswith("sk_") or c == "score_total"]
    grouped = df.groupby("ID_Institucion", as_index=False)[value_cols].mean(numeric_only=True)
    return grouped


def generar_excel_resultados(results: Dict) -> bytes:
    """
    Genera un Excel bonito con múltiples hojas de resultados
    
    Args:
        results: Diccionario con asignaciones, summary, util, metrics
    
    Returns:
        bytes: Contenido del Excel de descargar
    """
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Azul oscuro
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # ============= HOJA 1: ASIGNACIONES =============
    ws_asignaciones = wb.active
    ws_asignaciones.title = "Asignaciones"
    
    df_asignaciones = results["asignaciones"]
    
    # Escribir encabezados
    for col_idx, col_name in enumerate(df_asignaciones.columns, start=1):
        cell = ws_asignaciones.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
    
    # Escribir datos
    for row_idx, (_, row) in enumerate(df_asignaciones.iterrows(), start=2):
        for col_idx, val in enumerate(row, start=1):
            cell = ws_asignaciones.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.border = border
            
            # Formatear números
            if isinstance(val, (int, np.integer)) and col_idx == 7:  # Asignados
                cell.alignment = center_align
            elif isinstance(val, (float, np.floating)) and col_idx == 8:  # Score_unitario
                cell.value = round(val, 4)
                cell.number_format = '0.0000'
                cell.alignment = center_align
    
    # Ajustar anchos de columna
    ancho_default = 15
    anchos = {
        0: 13,  # ID_Institucion
        1: 25,  # Institucion
        2: 18,  # Programa
        3: 18,  # Tipo_Estudiante
        4: 22,  # Tipo_Practica
        5: 12,  # Semestre
        6: 12,  # Asignados
        7: 15,  # Score_unitario
    }
    
    for col_idx in range(1, len(df_asignaciones.columns) + 1):
        col_letter = ws_asignaciones.cell(row=1, column=col_idx).column_letter
        ws_asignaciones.column_dimensions[col_letter].width = anchos.get(col_idx - 1, ancho_default)
    
    # Congelar primera fila
    ws_asignaciones.freeze_panes = "A2"
    
    # ============= HOJA 2: RESUMEN =============
    ws_resumen = wb.create_sheet("Resumen")
    
    # Construir metrics a partir de los datos disponibles
    metrics = {
        "demanda_total": results.get("total_demanda", 0),
        "asignados": results.get("total_asignado", 0),
        "brecha": results.get("brecha", 0),
        "cobertura_pct": results.get("tasa_cobertura", 0) * 100,
        "num_instituciones": results["debug"].get("instituciones", 0),
        "pares_factibles": results["debug"].get("pares_factibles", 0),
        "pares_con_costo": results["debug"].get("pares_con_costo", 0),
        "criterios": results["debug"].get("criterios", 0)
    }
    
    # Títulos y valores
    resumen_data = [
        ("Métrica", "Valor"),
        ("Demanda Total (estudiantes)", metrics["demanda_total"]),
        ("Asignados", metrics["asignados"]),
        ("Brecha", metrics["brecha"]),
        ("Cobertura (%)", f"{metrics['cobertura_pct']:.2f}%"),
        ("Número de instituciones", metrics["num_instituciones"]),
        ("Pares factibles", metrics["pares_factibles"]),
        ("Pares con costo", metrics["pares_con_costo"]),
        ("Criterios activos", metrics["criterios"]),
    ]
    
    for row_idx, (metrica, valor) in enumerate(resumen_data, start=1):
        # Columna A - Métrica
        cell_a = ws_resumen.cell(row=row_idx, column=1)
        cell_a.value = metrica
        if row_idx == 1:
            cell_a.fill = header_fill
            cell_a.font = header_font
            cell_a.alignment = center_align
        else:
            cell_a.alignment = Alignment(horizontal="left", vertical="center")
        cell_a.border = border
        
        # Columna B - Valor
        cell_b = ws_resumen.cell(row=row_idx, column=2)
        cell_b.value = valor
        if row_idx == 1:
            cell_b.fill = header_fill
            cell_b.font = header_font
            cell_b.alignment = center_align
        else:
            cell_b.alignment = Alignment(horizontal="right", vertical="center")
        cell_b.border = border
    
    ws_resumen.column_dimensions["A"].width = 35
    ws_resumen.column_dimensions["B"].width = 20
    
    # ============= HOJA 3: UTILIZACIÓN =============
    ws_util = wb.create_sheet("Utilización")
    
    df_util = results["util"]
    
    if df_util is not None and not df_util.empty:
        # Escribir encabezados
        for col_idx, col_name in enumerate(df_util.columns, start=1):
            cell = ws_util.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align
        
        # Escribir datos
        for row_idx, (_, row) in enumerate(df_util.iterrows(), start=2):
            for col_idx, val in enumerate(row, start=1):
                cell = ws_util.cell(row=row_idx, column=col_idx)
                cell.value = val
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Formatear porcentaje
                col_name = df_util.columns[col_idx - 1]
                if col_name == "Utilización_%" and isinstance(val, (int, float)):
                    cell.value = round(val, 2)
                    cell.number_format = '0.00"%"'
        
        # Ajustar anchos
        ws_util.column_dimensions["A"].width = 13
        ws_util.column_dimensions["B"].width = 25
        ws_util.column_dimensions["C"].width = 18
        ws_util.column_dimensions["D"].width = 18
        ws_util.column_dimensions["E"].width = 15
    
    # ============= GUARDAR A BYTES =============
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generar_excel_refinado(results: Dict) -> bytes:
    """Genera Excel profesional multi-semestre con hoja de indicadores de alto impacto visual."""
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ---- Paleta corporativa ----
    COLOR_HEADER    = "0B3D5C"   # azul petróleo
    COLOR_BANNER    = "12557A"
    COLOR_ALT_ROW   = "EAF2F8"
    COLOR_AMARILLO  = "F4C430"   # baja ocupación
    COLOR_VERDE     = "27AE60"   # óptimo
    COLOR_ROJO      = "E74C3C"   # alerta
    COLOR_AMAR_BG   = "FDF2CC"
    COLOR_VERDE_BG  = "D5F0E0"
    COLOR_ROJO_BG   = "FAD7D2"
    COLOR_TRACK     = "ECF0F1"   # fondo de la barra
    WHITE           = "FFFFFF"
    GREY_TXT        = "5D6D7E"

    header_fill  = PatternFill("solid", fgColor=COLOR_HEADER)
    header_font  = Font(bold=True, color=WHITE, size=11)
    alt_fill     = PatternFill("solid", fgColor=COLOR_ALT_ROW)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_align   = Alignment(horizontal="left",   vertical="center")
    border       = Border(
        left=Side(style="thin", color="D5DBDB"), right=Side(style="thin", color="D5DBDB"),
        top=Side(style="thin", color="D5DBDB"),  bottom=Side(style="thin", color="D5DBDB"),
    )

    def write_header(ws, columns, row=1):
        for ci, name in enumerate(columns, 1):
            c = ws.cell(row=row, column=ci, value=name)
            c.fill, c.font, c.border, c.alignment = header_fill, header_font, border, center_align
        ws.freeze_panes = ws.cell(row=row + 1, column=1)

    def autofit(ws, columns, extra=4, start_row=1):
        for ci, name in enumerate(columns, 1):
            max_len = max(len(str(name)), 10)
            for row in ws.iter_rows(min_row=start_row + 1, min_col=ci, max_col=ci):
                for cell in row:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + extra, 55)

    df_asig = results["asignaciones"].copy()

    # ===========================================================
    # HOJA 1 — ASIGNACIONES
    # ===========================================================
    ws_a = wb.active
    ws_a.title = "Asignaciones"

    display_cols = ["Semestre", "Grupo_ID", "Tamano_Grupo", "Asignatura", "Set", "Rotacion",
                    "ID_Institucion", "Institucion", "Estudiantes", "Score_IPS"]
    display_cols = [c for c in display_cols if c in df_asig.columns]
    df_show = df_asig[display_cols]

    write_header(ws_a, display_cols)
    for ri, (_, row) in enumerate(df_show.iterrows(), start=2):
        fill_row = alt_fill if ri % 2 == 0 else None
        for ci, val in enumerate(row, start=1):
            cell = ws_a.cell(row=ri, column=ci, value=val)
            cell.border = border
            if fill_row:
                cell.fill = fill_row
            col_name = display_cols[ci - 1]
            if col_name in ("Semestre", "Grupo_ID", "Tamano_Grupo", "Estudiantes"):
                cell.alignment = center_align
            elif col_name == "Score_IPS" and isinstance(val, (float, np.floating)):
                cell.value = round(float(val), 4)
                cell.number_format = "0.0000"
                cell.alignment = center_align
            else:
                cell.alignment = left_align
    autofit(ws_a, display_cols)
    if df_show.shape[0] > 0:
        ws_a.auto_filter.ref = ws_a.dimensions

    # ===========================================================
    # HOJA 2 — RESUMEN
    # ===========================================================
    ws_r = wb.create_sheet("Resumen")
    por_sem = results.get("por_semestre", {})
    res_cols = ["Semestre", "Asignaturas", "Sets aplicados", "Estudiantes",
                "Asignados", "Grupos", "Tamaño grupos", "Calidad (score)"]
    write_header(ws_r, res_cols)
    ri = 2
    for sem in sorted(por_sem.keys()):
        d = por_sem[sem]
        sets_aplicados = ", ".join(sorted(set(d["sets"].values())))
        vals = [
            sem,
            ", ".join(d["asignaturas"]),
            sets_aplicados,
            d["n_estudiantes"],
            d["asignados"],
            d["n_grupos"],
            f"{d['min_group']}–{d['max_group']}",
            round(d["obj_value"], 4),
        ]
        for ci, val in enumerate(vals, start=1):
            cell = ws_r.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.alignment = center_align if ci != 2 and ci != 3 else left_align
            if ri % 2 == 0:
                cell.fill = alt_fill
        ri += 1
    autofit(ws_r, res_cols)
    if ri > 2:
        ws_r.auto_filter.ref = f"A1:{get_column_letter(len(res_cols))}{ri - 1}"

    # ===========================================================
    # HOJA 3 — INDICADORES_DEMANDA_OFERTA  (alto impacto visual)
    # ===========================================================
    ws_ind = wb.create_sheet("Indicadores_Demanda_Oferta")

    indicadores = results.get("indicadores", [])

    # --- Banner superior ---
    ws_ind.merge_cells("A1:H1")
    banner = ws_ind.cell(row=1, column=1, value="📊  INDICADORES DE DEMANDA vs. OFERTA")
    banner.fill = PatternFill("solid", fgColor=COLOR_BANNER)
    banner.font = Font(bold=True, color=WHITE, size=16)
    banner.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_ind.row_dimensions[1].height = 34

    ws_ind.merge_cells("A2:H2")
    sub = ws_ind.cell(
        row=2, column=1,
        value="🟡 Subutilización (<50%)   🟢 Óptimo (50–75%)   🟡 Alto (75–100%)   🔴 Alerta (>100%)",
    )
    sub.font = Font(italic=True, color=GREY_TXT, size=10)
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_ind.row_dimensions[2].height = 20

    # --- Encabezados (fila 4) ---
    BAR_SEGMENTS = 20  # nº de celdas que forman la barra visual
    header_row = 4
    ind_cols = ["Semestre", "Asignatura", "Set", "Demanda",
                "Oferta\nMáxima", "Ocupación\n(Dem/Oferta)", "Estado", "Interpretación"]
    # La barra ocupará columnas a la derecha (I en adelante)
    for ci, name in enumerate(ind_cols, 1):
        c = ws_ind.cell(row=header_row, column=ci, value=name)
        c.fill, c.font, c.border = header_fill, header_font, border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # Encabezado de la barra
    bar_start_col = len(ind_cols) + 1
    ws_ind.merge_cells(
        start_row=header_row, start_column=bar_start_col,
        end_row=header_row, end_column=bar_start_col + BAR_SEGMENTS - 1,
    )
    cbar = ws_ind.cell(row=header_row, column=bar_start_col, value="Indicador visual (0% ─────► 100%)")
    cbar.fill, cbar.font = header_fill, header_font
    cbar.alignment = Alignment(horizontal="center", vertical="center")
    ws_ind.row_dimensions[header_row].height = 30

    def estado_de(pct):
        if pct < 0.50:
            return "🟡 Subutilización", "Ocupación baja — hay holgura de oferta (<50%)", COLOR_AMARILLO, COLOR_AMAR_BG
        if pct <= 0.75:
            return "🟢 Óptimo", "Ocupación en el rango ideal (50–75%)", COLOR_VERDE, COLOR_VERDE_BG
        if pct <= 1.00:
            return "🟡 Alto", "Demanda cercana a la oferta máxima (75–100%)", COLOR_AMARILLO, COLOR_AMAR_BG
        return "🔴 Alerta", "Demanda supera la oferta disponible (>100%)", COLOR_ROJO, COLOR_ROJO_BG

    ri = header_row + 1
    for ind in indicadores:
        pct = float(ind["Pct_Demanda_Oferta"])
        estado, interp, color_bar, color_bg = estado_de(pct)

        vals = [
            ind["Semestre"], ind["Asignatura"], ind["Set"],
            ind["Demanda"], ind["Oferta_Maxima"], pct, estado, interp,
        ]
        for ci, val in enumerate(vals, start=1):
            cell = ws_ind.cell(row=ri, column=ci, value=val)
            cell.border = border
            col_name = ind_cols[ci - 1]
            if col_name.startswith("Ocupación"):
                cell.number_format = "0.0%"
                cell.alignment = center_align
                cell.font = Font(bold=True, color=color_bar)
                cell.fill = PatternFill("solid", fgColor=color_bg)
            elif col_name == "Estado":
                cell.alignment = center_align
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor=color_bg)
            elif col_name in ("Demanda", "Oferta\nMáxima", "Semestre"):
                cell.alignment = center_align
            else:
                cell.alignment = left_align

        # --- Barra visual con celdas coloreadas (data-bar manual) ---
        filled = int(round(min(pct, 1.0) * BAR_SEGMENTS))
        for seg in range(BAR_SEGMENTS):
            bc = ws_ind.cell(row=ri, column=bar_start_col + seg)
            if seg < filled:
                bc.fill = PatternFill("solid", fgColor=color_bar)
            else:
                bc.fill = PatternFill("solid", fgColor=COLOR_TRACK)
            bc.border = Border(top=Side(style="thin", color=WHITE), bottom=Side(style="thin", color=WHITE))
        # Etiqueta % al final de la barra
        lbl = ws_ind.cell(row=ri, column=bar_start_col + BAR_SEGMENTS, value=f"{pct*100:.0f}%")
        lbl.font = Font(bold=True, color=color_bar, size=10)
        lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws_ind.row_dimensions[ri].height = 22
        ri += 1

    # Anchos
    anchos_ind = {1: 11, 2: 30, 3: 24, 4: 11, 5: 11, 6: 15, 7: 18, 8: 38}
    for ci, w in anchos_ind.items():
        ws_ind.column_dimensions[get_column_letter(ci)].width = w
    for seg in range(BAR_SEGMENTS):
        ws_ind.column_dimensions[get_column_letter(bar_start_col + seg)].width = 2.6
    ws_ind.column_dimensions[get_column_letter(bar_start_col + BAR_SEGMENTS)].width = 7
    ws_ind.freeze_panes = ws_ind.cell(row=header_row + 1, column=1)
    if ri > header_row + 1:
        ws_ind.auto_filter.ref = f"A{header_row}:H{ri - 1}"

    # ===========================================================
    # GUARDAR
    # ===========================================================
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def procesar_datos(
    loader: DataLoader,
    set_id: str,
    semestre: str,
    total_estudiantes: int,
    programa_manual: str,
    tipo_est_manual: str,
    tipo_practica_manual: str
) -> Optional[Dict]:
    """Procesa datos y ejecuta optimización"""
    
    try:
        # Validar pesos
        loader.validate_pesas()
        
        # Obtener ponderaciones
        weights, crit_type = loader.get_ponderaciones_dict()

        # Validación explícita de suma de pesos activos
        pesos_activos = sum(float(v) for v in weights.values())
        if abs(pesos_activos - 1.0) > 1e-6:
            raise ValueError(
                f"Los pesos activos del set seleccionado deben sumar 1.0; suma actual={pesos_activos:.6f}"
            )
        
        st.write(f"✓ {len(weights)} criterios cargados")
        
        # Preparar cupos
        loader.cupos["Cupo_Estimado_Semestral"] = pd.to_numeric(
            loader.cupos["Cupo_Estimado_Semestral"], errors="coerce"
        ).fillna(0).astype(int)
        
        cupos_llenos = (loader.cupos["Cupo_Estimado_Semestral"] > 0).sum()
        
        # Si no hay cupos, generar ejemplo
        if cupos_llenos == 0:
            st.warning("⚠️ Plantilla sin cupos reales - usando datos de EJEMPLO")
            loader.cupos = generate_ejemplo_cupos(semestre)
            loader.costos = generate_ejemplo_costos(semestre)
        
        # Preparar costos
        loader.costos = loader.costos.copy()
        loader.costos["Cobro_EPP_num"] = loader.costos[
            "Cobro_EPP (No cobra/Cobra a la Universidad)"
        ].map({
            "No cobra EPP": 0,
            "Cobra EPP a la Universidad": 1,
        }).fillna(0)
        
        loader.costos["pct_contra"] = pd.to_numeric(
            loader.costos["%_Contraprestacion_Matricula (0-100)"], errors="coerce"
        )
        
        # Obtener demanda o usar demanda manual
        if loader.demanda is not None and not loader.demanda.empty and "Semestre" in loader.demanda.columns:
            demanda = loader.demanda[loader.demanda["Semestre"].astype(str) == str(semestre)].copy()
        else:
            st.info("ℹ️ Usando demanda manual definida en esta pantalla.")
            demanda = generate_ejemplo_demanda(
                semestre=semestre,
                total_estudiantes=total_estudiantes,
                programa=programa_manual,
                tipo_estudiante=tipo_est_manual,
                tipo_practica=tipo_practica_manual,
            )

        if demanda.empty:
            st.warning("⚠️ No se encontró demanda para el semestre seleccionado. Se aplicará demanda manual.")
            demanda = generate_ejemplo_demanda(
                semestre=semestre,
                total_estudiantes=total_estudiantes,
                programa=programa_manual,
                tipo_estudiante=tipo_est_manual,
                tipo_practica=tipo_practica_manual,
            )
        
        demanda["Demanda_Estudiantes"] = pd.to_numeric(
            demanda["Demanda_Estudiantes"], errors="coerce"
        ).fillna(0).astype(int)
        
        # Construir grupos y diccionarios
        groups = []
        for _, g in demanda.iterrows():
            groups.append((g["Programa"], g["Tipo_Estudiante"], g["Tipo_Practica"], g["Semestre"]))
        
        demand_dict = {
            (g["Programa"], g["Tipo_Estudiante"], g["Tipo_Practica"], g["Semestre"]): int(g["Demanda_Estudiantes"])
            for _, g in demanda.iterrows()
        }
        
        # Preparar cupos dict
        cap_dict = {}
        for _, r in loader.cupos.iterrows():
            inst_id = str(r["ID_Institucion"])
            prog = str(r["Programa"]).strip() if ("Programa" in loader.cupos.columns and pd.notna(r["Programa"])) else ""
            tipo_est = str(r["Tipo_Estudiante (Pregrado/Posgrado)"]).strip() if ("Tipo_Estudiante (Pregrado/Posgrado)" in loader.cupos.columns and pd.notna(r["Tipo_Estudiante (Pregrado/Posgrado)"])) else ""
            sem = str(r["Semestre (AAAA-S)"]).strip() if ("Semestre (AAAA-S)" in loader.cupos.columns and pd.notna(r["Semestre (AAAA-S)"])) else ""

            # fallback cuando vienen vacíos en el Excel
            if not prog:
                prog = programa_manual
            if not tipo_est:
                tipo_est = tipo_est_manual
            if not sem:
                sem = semestre

            cap_dict[(inst_id, prog, tipo_est, sem)] = int(r["Cupo_Estimado_Semestral"])
        
        # Instituciones
        instituciones = [str(j) for j in loader.oferta["ID_Institucion"].dropna().unique().tolist()]
        
        # Merge oferta + calidad (tolerante a columnas faltantes)
        base_raw = loader.oferta.merge(loader.calidad, on="ID_Institucion", how="left", suffixes=("", "_cal"))
        base = pd.DataFrame({"ID_Institucion": base_raw["ID_Institucion"]})

        rename_map = {
            "Acceso_Transporte_Publico (1-5)": "Acceso_Transporte_Publico",
            "MisionVisionProposito_AlineacionDocencia (1-5)": "MisionVisionProposito_AlineacionDocencia",
            "Evalua_Estudiantes_Profesores (0-5)": "Evalua_Estudiantes_Profesores",
            "Vinculacion_Planta_Especialistas_%": "Vinculacion_Planta_Especialistas_%",
            "Servicios_UCI (0/1)": "Servicios_UCI",
            "Servicios_UCIN (0/1)": "Servicios_UCIN",
            "Servicios_Pediatricos (0/1)": "Servicios_Pediatricos",
            "Servicios_Obstetricia (0/1)": "Servicios_Obstetricia",
            "Nro_Universidades_Comparten": "Nro_Universidades_Comparten",
        }

        for raw_col, norm_col in rename_map.items():
            if raw_col in base_raw.columns:
                base[norm_col] = base_raw[raw_col]

        # Compatibilidad: columna única UCI_UCIN
        if "Servicios_UCI_UCIN (0/1)" in base_raw.columns:
            base["Servicios_UCI_UCIN"] = base_raw["Servicios_UCI_UCIN (0/1)"]
        
        # Normalizar criterios
        S = ScoreCalculator.normalize_criteria(base)
        s_ids = S.index.astype(str)

        oferta_idx = loader.oferta.copy()
        oferta_idx["ID_Institucion"] = oferta_idx["ID_Institucion"].astype(str)
        oferta_idx = oferta_idx.set_index("ID_Institucion")

        calidad_idx = loader.calidad.copy()
        calidad_idx["ID_Institucion"] = calidad_idx["ID_Institucion"].astype(str)
        calidad_idx = calidad_idx.set_index("ID_Institucion")

        # Criterios adicionales del nuevo set
        if "Es_Hospital_Universitario" in loader.oferta.columns:
            hosp_uni = oferta_idx["Es_Hospital_Universitario"].reindex(s_ids, fill_value=0)
            S["Es_Hospital_Universitario_norm"] = hosp_uni.apply(to_bool01).values

        if "Escenario_Avalado_Practicas" in loader.oferta.columns:
            esc = oferta_idx["Escenario_Avalado_Practicas"].reindex(s_ids, fill_value=0)
            S["Escenario_Avalado_Practicas_norm"] = esc.apply(to_bool01).values

        # Compatibilidad UCI/UCIN/UCI_UCIN
        if "Servicios_UCI_UCIN" in base.columns:
            combo = base.set_index("ID_Institucion")["Servicios_UCI_UCIN"].reindex(s_ids, fill_value=0)
            S["Servicios_UCI_UCIN_norm"] = combo.apply(to_bool01).astype(float).values
        elif "Servicios_UCI" in base.columns or "Servicios_UCIN" in base.columns:
            uci = base.set_index("ID_Institucion").get("Servicios_UCI", pd.Series(index=s_ids, data=0)).reindex(s_ids, fill_value=0)
            ucin = base.set_index("ID_Institucion").get("Servicios_UCIN", pd.Series(index=s_ids, data=0)).reindex(s_ids, fill_value=0)
            S["Servicios_UCI_UCIN_norm"] = np.maximum(
                uci.apply(to_bool01).astype(float),
                ucin.apply(to_bool01).astype(float),
            )

        if "Admiten_Docentes_Externos (Sí/No)" in loader.calidad.columns:
            adm = calidad_idx["Admiten_Docentes_Externos (Sí/No)"].reindex(s_ids, fill_value="No")
            adm_num = adm.astype(str).str.strip().str.lower().map({"sí": 1, "si": 1, "yes": 1, "1": 1, "true": 1}).fillna(0)
            S["Admiten_Docentes_Externos_norm"] = adm_num.values

        for col_raw, col_norm in [
            ("Areas_Bienestar (0/1)", "Areas_Bienestar_norm"),
            ("Areas_Academicas (0/1)", "Areas_Academicas_norm"),
        ]:
            if col_raw in loader.calidad.columns:
                col_s = calidad_idx[col_raw].reindex(s_ids, fill_value=0)
                S[col_norm] = pd.to_numeric(col_s, errors="coerce").fillna(0).clip(0, 1).values
        
        # Calcular scores V(j,g)
        def lookup_costo(id_inst, prog, tipo_est, tipo_pract, semestre):
            id_inst = str(id_inst)
            prog = str(prog)
            tipo_est = str(tipo_est)
            tipo_pract = str(tipo_pract)
            semestre = str(semestre)

            c = loader.costos.copy()
            c["ID_Institucion"] = c["ID_Institucion"].astype(str)
            if "Tipo_Estudiante_Costo" in c.columns:
                c["Tipo_Estudiante_Costo"] = c["Tipo_Estudiante_Costo"].astype(str)
            if "Tipo_Practica_Costo" in c.columns:
                c["Tipo_Practica_Costo"] = c["Tipo_Practica_Costo"].astype(str)
            if "Semestre_Vigencia (AAAA-S)" in c.columns:
                c["Semestre_Vigencia (AAAA-S)"] = c["Semestre_Vigencia (AAAA-S)"].astype(str)
            if "Programa_Costo" in c.columns:
                c["Programa_Costo"] = c["Programa_Costo"].astype(str)
            if "EPP_Exigidos_num" not in c.columns and "EPP_Exigidos (Sin exigencia/Parcial/Completo + detalle)" in c.columns:
                c["EPP_Exigidos_num"] = c["EPP_Exigidos (Sin exigencia/Parcial/Completo + detalle)"].apply(map_epp_exigidos)

            # 1) Match estricto
            df = c[c["ID_Institucion"] == id_inst]
            if "Tipo_Estudiante_Costo" in c.columns:
                df = df[df["Tipo_Estudiante_Costo"] == tipo_est]
            if "Tipo_Practica_Costo" in c.columns:
                df = df[df["Tipo_Practica_Costo"] == tipo_pract]
            if "Semestre_Vigencia (AAAA-S)" in c.columns:
                df = df[df["Semestre_Vigencia (AAAA-S)"] == semestre]

            if "Programa_Costo" in c.columns:
                df1 = df[df["Programa_Costo"] == prog]
                df2 = df[df["Programa_Costo"] == "Todos"]
            else:
                df1 = df
                df2 = pd.DataFrame()

            if len(df1) > 0:
                row = df1.iloc[0]
                return row["pct_contra"], row["Cobro_EPP_num"], row.get("EPP_Exigidos_num", np.nan)
            if len(df2) > 0:
                row = df2.iloc[0]
                return row["pct_contra"], row["Cobro_EPP_num"], row.get("EPP_Exigidos_num", np.nan)

            # 2) Fallback sin tipo_practica
            df = c[c["ID_Institucion"] == id_inst]
            if "Tipo_Estudiante_Costo" in c.columns:
                df = df[df["Tipo_Estudiante_Costo"] == tipo_est]
            if "Semestre_Vigencia (AAAA-S)" in c.columns:
                df = df[df["Semestre_Vigencia (AAAA-S)"] == semestre]

            if "Programa_Costo" in c.columns:
                df1 = df[df["Programa_Costo"] == prog]
                df2 = df[df["Programa_Costo"] == "Todos"]
            else:
                df1 = df
                df2 = pd.DataFrame()

            if len(df1) > 0:
                row = df1.iloc[0]
                return row["pct_contra"], row["Cobro_EPP_num"], row.get("EPP_Exigidos_num", np.nan)
            if len(df2) > 0:
                row = df2.iloc[0]
                return row["pct_contra"], row["Cobro_EPP_num"], row.get("EPP_Exigidos_num", np.nan)

            # 3) Fallback por institución (neutral si falta)
            df = c[c["ID_Institucion"] == id_inst]
            if len(df) > 0:
                row = df.iloc[0]
                pct = row["pct_contra"] if pd.notna(row["pct_contra"]) else 50.0
                epp = row["Cobro_EPP_num"] if pd.notna(row["Cobro_EPP_num"]) else 0
                epp_exig = row.get("EPP_Exigidos_num", np.nan)
                return pct, epp, epp_exig

            return 50.0, 0, np.nan
        
        # Normalizar llaves de criterios de forma robusta
        weights_norm = {}
        for k, w in weights.items():
            key = clean_criterio_codigo(k)
            weights_norm[key] = weights_norm.get(key, 0.0) + float(w)

        # Validación explícita de pesos limpios
        pesos_limpios_sum = sum(weights_norm.values())
        if abs(pesos_limpios_sum - 1.0) > 1e-6:
            raise ValueError(
                f"La suma de pesos activos (limpios) debe ser 1.0; suma actual={pesos_limpios_sum:.6f}"
            )

        weights_raw_df = pd.DataFrame(
            {
                "criterio_raw": list(weights.keys()),
                "criterio_clean": [clean_criterio_codigo(k) for k in weights.keys()],
                "peso": list(weights.values()),
                "tipo": [crit_type.get(k) for k in weights.keys()],
            }
        )
        weights_clean_df = weights_raw_df.groupby("criterio_clean", as_index=False)["peso"].sum()

        criteria_status_rows = []
        for k in weights_norm.keys():
            source = "S_norm"
            if k == "%_Contraprestacion_Matricula":
                source = "costo_pct"
            elif k == "Cobro_EPP":
                source = "costo_cobro_epp"
            elif k == "EPP_Exigidos":
                source = "costo_epp_exigidos"
            elif k == "Admiten_Docentes_Externos":
                source = "calidad_bool"
            elif f"{k}_norm" not in S.columns:
                source = "missing"
            criteria_status_rows.append({"criterio": k, "source": source})
        criteria_status_df = pd.DataFrame(criteria_status_rows)

        special_criteria = {
            "%_Contraprestacion_Matricula",
            "Cobro_EPP",
            "EPP_Exigidos",
            "Admiten_Docentes_Externos",
        }

        missing_criteria = set()
        for k in weights_norm.keys():
            if k in special_criteria:
                continue
            if f"{k}_norm" not in S.columns:
                missing_criteria.add(k)

        V = {}
        count_factible = 0
        count_asignado = 0
        pair_score_rows = []
        epp_fallback_pairs = 0
        
        for j in instituciones:
            for g in groups:
                cap = cap_dict.get((j, g[0], g[1], g[3]), 0)
                if cap <= 0:
                    continue
                
                count_factible += 1
                p, n, t, s = g
                
                pct_contra, cobro_epp, epp_exig_val = lookup_costo(j, p, n, t, s)
                
                if pd.isna(pct_contra):
                    continue
                
                count_asignado += 1
                
                contra_norm = 1.0 - float(pct_contra) / 100.0
                cobro_epp_norm = 1.0 - float(cobro_epp)
                if pd.notna(epp_exig_val):
                    epp_exig_norm = 1.0 - float(epp_exig_val)
                else:
                    epp_exig_norm = cobro_epp_norm
                    epp_fallback_pairs += 1
                
                score = 0.0
                score_row = {"ID_Institucion": j}
                for k, w in weights_norm.items():
                    if w <= 0:
                        continue
                    
                    if k == "%_Contraprestacion_Matricula":
                        sk = contra_norm
                    elif k == "Cobro_EPP":
                        sk = cobro_epp_norm
                    elif k == "EPP_Exigidos":
                        sk = epp_exig_norm
                    elif k == "Admiten_Docentes_Externos":
                        sk = S.loc[j, "Admiten_Docentes_Externos_norm"] if "Admiten_Docentes_Externos_norm" in S.columns else 0.0
                    else:
                        sk = S.loc[j, f"{k}_norm"] if f"{k}_norm" in S.columns else 0.0
                        if f"{k}_norm" not in S.columns:
                            missing_criteria.add(k)
                        if pd.isna(sk):
                            sk = 0.0
                    
                    score += w * float(sk)
                    score_row[f"sk_{k}"] = float(sk)
                
                V[(j, g)] = score
                score_row["score_total"] = float(score)
                pair_score_rows.append(score_row)

        score_debug_df = compute_scores_debug(pair_score_rows)
        
        st.write(f"✓ Pares (j,g) para optimización: {len(V)}")
        
        # Optimizar
        optimizer = Optimizer(verbose=False)
        results_df = optimizer.optimize(V, demand_dict, cap_dict, instituciones, groups, semestre)

        # Agregar nombre de institución a resultados
        if not results_df.empty and "Institucion" in loader.oferta.columns:
            oferta_names = loader.oferta[["ID_Institucion", "Institucion"]].copy()
            oferta_names["ID_Institucion"] = oferta_names["ID_Institucion"].astype(str)
            results_df["ID_Institucion"] = results_df["ID_Institucion"].astype(str)
            results_df = results_df.merge(oferta_names, on="ID_Institucion", how="left")
            # Orden de columnas más amigable
            cols = [
                "ID_Institucion", "Institucion", "Programa", "Tipo_Estudiante",
                "Tipo_Practica", "Semestre", "Asignados", "Score_unitario"
            ]
            results_df = results_df[[c for c in cols if c in results_df.columns]]

        if not results_df.empty and "Score_unitario" in results_df.columns:
            results_df["Score_unitario"] = pd.to_numeric(results_df["Score_unitario"], errors="coerce").round(4)
        
        obj_val = optimizer.get_objective_value()

        score_consistency = {
            "max_abs_diff": None,
            "mean_abs_diff": None,
        }
        if not results_df.empty and not score_debug_df.empty:
            compare = results_df[["ID_Institucion", "Score_unitario"]].copy()
            compare["ID_Institucion"] = compare["ID_Institucion"].astype(str)
            score_debug_df["ID_Institucion"] = score_debug_df["ID_Institucion"].astype(str)
            compare = compare.merge(
                score_debug_df[["ID_Institucion", "score_total"]],
                on="ID_Institucion",
                how="left",
            )
            compare["abs_diff"] = (pd.to_numeric(compare["Score_unitario"], errors="coerce") - pd.to_numeric(compare["score_total"], errors="coerce")).abs()
            score_consistency["max_abs_diff"] = float(compare["abs_diff"].max()) if compare["abs_diff"].notna().any() else None
            score_consistency["mean_abs_diff"] = float(compare["abs_diff"].mean()) if compare["abs_diff"].notna().any() else None
        
        # Compilar resultados
        total_demanda = sum(demand_dict.values())
        total_asignado = results_df["Asignados"].sum() if not results_df.empty else 0
        brecha = total_demanda - total_asignado
        tasa_cobertura = (total_asignado / total_demanda * 100) if total_demanda > 0 else 0
        
        # Resumen por grupo
        if not results_df.empty:
            summary = results_df.groupby(["Programa", "Tipo_Estudiante", "Tipo_Practica"])["Asignados"].sum().reset_index()
            summary.columns = ["Programa", "Tipo_Estudiante", "Tipo_Practica", "Asignados"]
            summary["Demanda"] = summary.apply(
                lambda row: demand_dict.get((row["Programa"], row["Tipo_Estudiante"], row["Tipo_Practica"], semestre), 0),
                axis=1
            )
            summary["Gap"] = summary["Demanda"] - summary["Asignados"]
        else:
            summary = pd.DataFrame()
        
        # Utilización
        if not results_df.empty:
            util = results_df.groupby("ID_Institucion")["Asignados"].sum().reset_index()
            util.columns = ["ID_Institucion", "Estudiantes_Asignados"]
            util["Capacidad"] = util["ID_Institucion"].apply(
                lambda j: sum(v for k, v in cap_dict.items() if k[0] == j)
            )
            util["Capacidad"] = pd.to_numeric(util["Capacidad"], errors="coerce").fillna(0)
            util["Utilización_%"] = np.where(
                util["Capacidad"] > 0,
                (util["Estudiantes_Asignados"] / util["Capacidad"] * 100).round(1),
                0.0,
            )
            if "Institucion" in results_df.columns:
                util = util.merge(
                    results_df[["ID_Institucion", "Institucion"]].drop_duplicates(),
                    on="ID_Institucion",
                    how="left",
                )
        else:
            util = pd.DataFrame()

        base_debug_cols = [
            "ID_Institucion",
            "Acceso_Transporte_Publico (1-5)",
            "MisionVisionProposito_AlineacionDocencia (1-5)",
            "Evalua_Estudiantes_Profesores (0-5)",
            "Vinculacion_Planta_Especialistas_%",
            "Servicios_UCI (0/1)",
            "Servicios_UCIN (0/1)",
            "Servicios_UCI_UCIN (0/1)",
            "Servicios_Pediatricos (0/1)",
            "Servicios_Obstetricia (0/1)",
            "Nro_Universidades_Comparten",
            "Es_Hospital_Universitario",
            "Escenario_Avalado_Practicas",
            "Admiten_Docentes_Externos (Sí/No)",
            "Areas_Bienestar (0/1)",
            "Areas_Academicas (0/1)",
        ]
        base_debug = base_raw[[c for c in base_debug_cols if c in base_raw.columns]].copy()

        costos_debug_cols = [
            "ID_Institucion",
            "Programa_Costo",
            "Tipo_Estudiante_Costo",
            "Tipo_Practica_Costo",
            "Semestre_Vigencia (AAAA-S)",
            "%_Contraprestacion_Matricula (0-100)",
            "EPP_Exigidos (Sin exigencia/Parcial/Completo + detalle)",
            "Cobro_EPP (No cobra/Cobra a la Universidad)",
        ]
        costos_debug = loader.costos[[c for c in costos_debug_cols if c in loader.costos.columns]].copy()
        
        return {
            "asignaciones": results_df,
            "summary": summary,
            "util": util,
            "score_debug": score_debug_df,
            "total_demanda": total_demanda,
            "total_asignado": total_asignado,
            "brecha": brecha,
            "tasa_cobertura": tasa_cobertura,
            "obj_value": obj_val,
            "debug": {
                "instituciones": len(instituciones),
                "grupos": len(groups),
                "pares_factibles": count_factible,
                "pares_con_costo": count_asignado,
                "criterios": len(weights_norm),
                "missing_criteria": sorted(list(missing_criteria)),
                "epp_exigidos_fallback_pairs": epp_fallback_pairs,
                "score_consistency": score_consistency,
                "weights_raw": weights_raw_df,
                "weights_clean": weights_clean_df,
                "criteria_status": criteria_status_df,
                "base_debug": base_debug,
                "costos_debug": costos_debug,
                "score_debug": score_debug_df,
                "instituciones_list": instituciones,
                "groups_list": groups,
            }
        }
    
    except Exception as e:
        logger.error(f"Error procesando datos: {e}")
        st.error(f"❌ Error: {str(e)}")
        return None


def _prepare_score_matrix(loader: DataLoader) -> pd.DataFrame:
    """Construye la matriz de criterios normalizados S (indexada por ID_Institucion str).

    Centraliza toda la lógica de renombrado/normalización de criterios que antes
    estaba duplicada. También prepara columnas auxiliares en loader.costos.
    """
    base_raw = loader.oferta.merge(loader.calidad, on="ID_Institucion", how="left", suffixes=("", "_cal"))
    base = pd.DataFrame({"ID_Institucion": base_raw["ID_Institucion"]})

    rename_map = {
        "Acceso_Transporte_Publico (1-5)": "Acceso_Transporte_Publico",
        "MisionVisionProposito_AlineacionDocencia (1-5)": "MisionVisionProposito_AlineacionDocencia",
        "Evalua_Estudiantes_Profesores (0-5)": "Evalua_Estudiantes_Profesores",
        "Vinculacion_Planta_Especialistas_%": "Vinculacion_Planta_Especialistas_%",
        "Servicios_UCI (0/1)": "Servicios_UCI",
        "Servicios_UCIN (0/1)": "Servicios_UCIN",
        "Servicios_Pediatricos (0/1)": "Servicios_Pediatricos",
        "Servicios_Obstetricia (0/1)": "Servicios_Obstetricia",
        "Nro_Universidades_Comparten": "Nro_Universidades_Comparten",
    }
    for raw_col, norm_col in rename_map.items():
        if raw_col in base_raw.columns:
            base[norm_col] = base_raw[raw_col]

    S = ScoreCalculator.normalize_criteria(base)
    s_ids = S.index.astype(str)

    oferta_idx = loader.oferta.copy()
    oferta_idx["ID_Institucion"] = oferta_idx["ID_Institucion"].astype(str)
    oferta_idx = oferta_idx.set_index("ID_Institucion")

    calidad_idx = loader.calidad.copy()
    calidad_idx["ID_Institucion"] = calidad_idx["ID_Institucion"].astype(str)
    calidad_idx = calidad_idx.set_index("ID_Institucion")

    if "Es_Hospital_Universitario" in loader.oferta.columns:
        S["Es_Hospital_Universitario_norm"] = oferta_idx["Es_Hospital_Universitario"].reindex(s_ids, fill_value=0).apply(to_bool01).values

    if "Escenario_Avalado_Practicas" in loader.oferta.columns:
        S["Escenario_Avalado_Practicas_norm"] = oferta_idx["Escenario_Avalado_Practicas"].reindex(s_ids, fill_value=0).apply(to_bool01).values

    if "Servicios_UCI_UCIN (0/1)" in base_raw.columns:
        combo = base.set_index("ID_Institucion")["Servicios_UCI_UCIN (0/1)"].reindex(s_ids, fill_value=0)
        S["Servicios_UCI_UCIN_norm"] = combo.apply(to_bool01).astype(float).values
    elif "Servicios_UCI" in base.columns or "Servicios_UCIN" in base.columns:
        uci = base.set_index("ID_Institucion").get("Servicios_UCI", pd.Series(index=s_ids, data=0)).reindex(s_ids, fill_value=0)
        ucin = base.set_index("ID_Institucion").get("Servicios_UCIN", pd.Series(index=s_ids, data=0)).reindex(s_ids, fill_value=0)
        S["Servicios_UCI_UCIN_norm"] = np.maximum(
            uci.apply(to_bool01).astype(float),
            ucin.apply(to_bool01).astype(float),
        )

    if "Admiten_Docentes_Externos (Sí/No)" in loader.calidad.columns:
        adm = calidad_idx["Admiten_Docentes_Externos (Sí/No)"].reindex(s_ids, fill_value="No")
        adm_num = adm.astype(str).str.strip().str.lower().map({"sí": 1, "si": 1, "yes": 1, "1": 1, "true": 1}).fillna(0)
        S["Admiten_Docentes_Externos_norm"] = adm_num.values

    for col_raw, col_norm in [
        ("Areas_Bienestar (0/1)", "Areas_Bienestar_norm"),
        ("Areas_Academicas (0/1)", "Areas_Academicas_norm"),
    ]:
        if col_raw in loader.calidad.columns:
            col_s = calidad_idx[col_raw].reindex(s_ids, fill_value=0)
            S[col_norm] = pd.to_numeric(col_s, errors="coerce").fillna(0).clip(0, 1).values

    loader.costos = loader.costos.copy()
    loader.costos["Cobro_EPP_num"] = loader.costos[
        "Cobro_EPP (No cobra/Cobra a la Universidad)"
    ].map({"No cobra EPP": 0, "Cobra EPP a la Universidad": 1}).fillna(0)
    loader.costos["pct_contra"] = pd.to_numeric(
        loader.costos["%_Contraprestacion_Matricula (0-100)"], errors="coerce"
    )
    if "EPP_Exigidos (Sin exigencia/Parcial/Completo + detalle)" in loader.costos.columns:
        loader.costos["EPP_Exigidos_num"] = loader.costos[
            "EPP_Exigidos (Sin exigencia/Parcial/Completo + detalle)"
        ].apply(map_epp_exigidos)

    return S


def _weights_norm_for_set(loader: DataLoader, set_id: str) -> dict:
    """Obtiene y valida los pesos de un set, devolviéndolos limpios y normalizados."""
    loader.validate_pesas(set_id=set_id)
    weights, _ = loader.get_ponderaciones_dict(set_id=set_id)
    pesos_activos = sum(float(v) for v in weights.values())
    if abs(pesos_activos - 1.0) > 1e-6:
        raise ValueError(
            f"Los pesos del set '{set_id}' deben sumar 1.0; suma actual={pesos_activos:.6f}"
        )
    weights_norm = {}
    for k, w in weights.items():
        key = clean_criterio_codigo(k)
        weights_norm[key] = weights_norm.get(key, 0.0) + float(w)
    return weights_norm


def _scores_for_set(loader: DataLoader, S: pd.DataFrame, weights_norm: dict, ips_ids: set) -> dict:
    """Calcula {j: score} para las IPS dadas usando weights_norm sobre la matriz S."""
    s_ids = S.index.astype(str)
    costos = loader.costos.copy()
    costos["ID_Institucion"] = costos["ID_Institucion"].astype(str)
    costos_by_id = {jid: grp for jid, grp in costos.groupby("ID_Institucion")}

    scores = {}
    for j in ips_ids:
        if j not in set(s_ids):
            scores[j] = 0.0
            continue
        df_c = costos_by_id.get(j)
        score = 0.0
        for k, w in weights_norm.items():
            if w <= 0:
                continue
            if k == "%_Contraprestacion_Matricula":
                if df_c is not None and len(df_c) > 0:
                    pct = df_c["pct_contra"].iloc[0]
                    sk = 1.0 - float(pct) / 100.0 if pd.notna(pct) else 0.5
                else:
                    sk = 0.5
            elif k == "Cobro_EPP":
                if df_c is not None and len(df_c) > 0:
                    sk = 1.0 - float(df_c["Cobro_EPP_num"].iloc[0])
                else:
                    sk = 1.0
            elif k == "EPP_Exigidos":
                if df_c is not None and len(df_c) > 0 and "EPP_Exigidos_num" in df_c.columns:
                    val = df_c["EPP_Exigidos_num"].iloc[0]
                    sk = 1.0 - float(val) if pd.notna(val) else 0.5
                else:
                    sk = 0.5
            elif k == "Admiten_Docentes_Externos":
                sk = S.loc[j, "Admiten_Docentes_Externos_norm"] if "Admiten_Docentes_Externos_norm" in S.columns else 0.0
            else:
                col = f"{k}_norm"
                sk = S.loc[j, col] if col in S.columns else 0.0
                if pd.isna(sk):
                    sk = 0.0
            score += w * float(sk)
        scores[j] = round(score, 4)
    return scores


def procesar_refinado(
    loader: DataLoader,
    selecciones: list,
    n_por_semestre: dict,
    semestre_vigencia: str,
) -> Optional[Dict]:
    """Optimización refinada multi-semestre con un set de ponderaciones por asignatura.

    Args:
        selecciones: lista de dicts {"semestre": int, "asignatura": str, "set_id": str}.
        n_por_semestre: {semestre: n_estudiantes}.
    """
    try:
        if not selecciones:
            st.error("❌ No hay asignaturas seleccionadas para optimizar.")
            return None

        OFERTA_MAXIMA = 75

        # Preparar matriz de criterios una sola vez (común a todos los sets)
        S = _prepare_score_matrix(loader)

        # Agrupar selecciones por semestre (los estudiantes difieren por semestre)
        por_sem = {}
        for sel in selecciones:
            por_sem.setdefault(int(sel["semestre"]), {})[str(sel["asignatura"])] = str(sel["set_id"])

        # Cache de scores por set para no recomputar
        scores_cache: Dict[str, dict] = {}

        combined_rows = []
        por_semestre_detalle = {}
        indicadores = []
        scores_aj_global = {}
        obj_total = 0.0

        oferta_names = None
        if "Institucion" in loader.oferta.columns:
            oferta_names = loader.oferta[["ID_Institucion", "Institucion"]].copy()
            oferta_names["ID_Institucion"] = oferta_names["ID_Institucion"].astype(str)

        for sem in sorted(por_sem.keys()):
            set_by_asig = por_sem[sem]
            asigs = sorted(set_by_asig.keys())

            n_estudiantes = int(n_por_semestre.get(sem, 0) or 0)
            constraints = get_group_constraints(sem)
            min_g, max_g = constraints["min"], constraints["max"]

            st.markdown(f"**▶ Semestre {sem}** — asignaturas: {asigs} · estudiantes: {n_estudiantes}")

            if n_estudiantes < min_g:
                st.error(f"Semestre {sem}: se necesitan al menos {min_g} estudiantes para formar un grupo. Omitido.")
                continue

            cap_dict = loader.get_rotaciones_dict(sem, asigs)
            ar_dict = loader.get_asignaturas_rotaciones(sem, asigs)

            if not ar_dict:
                st.warning(f"⚠️ Semestre {sem}: sin rotaciones válidas para las asignaturas seleccionadas. Omitido.")
                continue

            # Rotaciones sin IPS
            for asig, rots in ar_dict.items():
                for rot in rots:
                    if not any(a == asig and r == rot for (a, r, j) in cap_dict):
                        st.warning(f"⚠️ Sem {sem} · {asig} / {rot}: sin IPS con cupo disponible.")

            ips_sem = {j for (a, r, j) in cap_dict}

            # Score por (asignatura, IPS) según el set de cada asignatura
            scores_aj = {}
            for asig in asigs:
                sid = set_by_asig[asig]
                if sid not in scores_cache:
                    weights_norm = _weights_norm_for_set(loader, sid)
                    scores_cache[sid] = _scores_for_set(loader, S, weights_norm, ips_sem)
                else:
                    # asegurar que cubra estas IPS
                    faltantes = ips_sem - set(scores_cache[sid].keys())
                    if faltantes:
                        weights_norm = _weights_norm_for_set(loader, sid)
                        scores_cache[sid].update(_scores_for_set(loader, S, weights_norm, faltantes))
                ips_scores = scores_cache[sid]
                for j in ips_sem:
                    scores_aj[(asig, j)] = ips_scores.get(j, 0.0)

            scores_aj_global.update(scores_aj)

            optimizer = GroupOptimizer(verbose=False)
            res_df = optimizer.optimize(
                scores=scores_aj,
                cap_dict=cap_dict,
                asignaturas_rotaciones=ar_dict,
                n_estudiantes=n_estudiantes,
                min_group=min_g,
                max_group=max_g,
            )

            if res_df is None or res_df.empty:
                st.warning(f"⚠️ Semestre {sem}: el optimizador no encontró asignaciones factibles.")
                continue

            # Enriquecer
            res_df["Semestre"] = sem
            res_df["Set"] = res_df["Asignatura"].map(set_by_asig)
            res_df["ID_Institucion"] = res_df["ID_Institucion"].astype(str)
            if oferta_names is not None:
                res_df = res_df.merge(oferta_names, on="ID_Institucion", how="left")
            # Grupo etiquetado por semestre para que sea único en la salida combinada
            res_df["Grupo_ID"] = res_df["Grupo"].map(lambda g: f"S{sem}-G{g}")

            combined_rows.append(res_df)
            obj_total += float(optimizer.get_objective_value() or 0.0)

            n_grupos = res_df["Grupo"].nunique()
            asignados_sem = int(res_df.groupby("Grupo")["Tamano_Grupo"].first().sum())
            por_semestre_detalle[sem] = {
                "n_estudiantes": n_estudiantes,
                "asignados": asignados_sem,
                "n_grupos": int(n_grupos),
                "min_group": min_g,
                "max_group": max_g,
                "asignaturas": asigs,
                "sets": set_by_asig,
                "obj_value": float(optimizer.get_objective_value() or 0.0),
            }

            # Indicadores por (semestre, asignatura)
            for asig in asigs:
                pct = n_estudiantes / OFERTA_MAXIMA if OFERTA_MAXIMA > 0 else 0.0
                indicadores.append({
                    "Semestre": sem,
                    "Asignatura": asig,
                    "Set": set_by_asig[asig],
                    "Demanda": n_estudiantes,
                    "Asignados": asignados_sem,
                    "Oferta_Maxima": OFERTA_MAXIMA,
                    "Pct_Demanda_Oferta": round(pct, 4),
                })

        if not combined_rows:
            st.error("❌ No se obtuvieron asignaciones en ningún semestre seleccionado.")
            return None

        df_all = pd.concat(combined_rows, ignore_index=True)

        total_estudiantes = sum(int(n_por_semestre.get(s, 0) or 0) for s in por_sem.keys())
        total_asignado = sum(d["asignados"] for d in por_semestre_detalle.values())
        n_grupos_total = sum(d["n_grupos"] for d in por_semestre_detalle.values())

        # Scores planos por IPS (promedio entre asignaturas) para gráficos avanzados
        scores_flat = {}
        tmp = {}
        for (a, j), v in scores_aj_global.items():
            tmp.setdefault(j, []).append(v)
        for j, vals in tmp.items():
            scores_flat[j] = round(sum(vals) / len(vals), 4)

        return {
            "modo": "refinado_multi",
            "asignaciones": df_all,
            "por_semestre": por_semestre_detalle,
            "indicadores": indicadores,
            "total_estudiantes": total_estudiantes,
            "total_asignado": total_asignado,
            "n_grupos_total": n_grupos_total,
            "obj_value": obj_total,
            "scores": scores_flat,
            "scores_aj": scores_aj_global,
            "selecciones": selecciones,
        }

    except Exception as e:
        logger.error(f"Error procesando refinado: {e}")
        st.error(f"❌ Error: {str(e)}")
        return None



def _render_heatmap_grupo_ips(df_asig, key_suffix=""):
    import plotly.express as px
    if "Institucion" not in df_asig.columns or "Score_IPS" not in df_asig.columns:
        return
    if df_asig["Score_IPS"].nunique() <= 1:
        return
    grupo_col = "Grupo_ID" if "Grupo_ID" in df_asig.columns else "Grupo"
    pivot = df_asig.pivot_table(
        values="Score_IPS", index=grupo_col, columns="Institucion", aggfunc="mean"
    )
    fig = px.imshow(
        pivot,
        color_continuous_scale="RdYlGn",
        aspect="auto",
        zmin=0, zmax=1,
        text_auto=".2f",
        title="Score promedio por Grupo e IPS",
    )
    fig.update_xaxes(tickangle=45)
    st.plotly_chart(fig, use_container_width=True, key=f"heatmap_score_{key_suffix}")
    st.caption("Cada celda muestra el score promedio de la IPS para un grupo. Verde = alto, rojo = bajo. Sirve para detectar si un grupo está cayendo sistemáticamente en IPS de baja calidad.")


def _render_quadrant_calidad_costo(df_asig, results, loader, key_suffix=""):
    import plotly.express as px
    if "scores" not in results or not results["scores"]:
        return

    costos = loader.costos.copy()
    if "ID_Institucion" in costos.columns:
        costos["ID_Institucion"] = costos["ID_Institucion"].astype(str)

    if "%_Contraprestacion_Matricula (0-100)" not in costos.columns:
        return

    oferta_idx = loader.oferta.copy()
    if "ID_Institucion" in oferta_idx.columns:
        oferta_idx["ID_Institucion"] = oferta_idx["ID_Institucion"].astype(str)

    rows = []
    for j, score in results["scores"].items():
        df_c = costos[costos["ID_Institucion"] == j]
        if len(df_c) > 0:
            pct = df_c["%_Contraprestacion_Matricula (0-100)"].iloc[0]
            pct = float(pct) if pd.notna(pct) else 50.0
        else:
            pct = 50.0
        inst_match = oferta_idx[oferta_idx["ID_Institucion"] == j]
        name = inst_match["Institucion"].iloc[0][:30] if len(inst_match) > 0 else j
        used = df_asig[df_asig["ID_Institucion"].astype(str) == j]["Estudiantes"].sum()
        rows.append({"ID_IPS": j, "IPS": name, "Score": score, "Costo_%": pct, "Usados": used})

    df_q = pd.DataFrame(rows)
    if df_q.empty or df_q["Score"].nunique() <= 1:
        return

    fig = px.scatter(
        df_q, x="Costo_%", y="Score", size="Usados", color="Score",
        hover_name="IPS", text="IPS",
        color_continuous_scale="RdYlGn", size_max=40,
        title="Frontera Eficiencia: Score vs Costo (tamaño = cupos usados)",
    )
    fig.update_traces(textposition="top center", textfont_size=8)
    fig.add_vline(x=df_q["Costo_%"].median(), line_dash="dot", line_color="gray", opacity=0.5)
    fig.add_hline(y=df_q["Score"].median(), line_dash="dot", line_color="gray", opacity=0.5)
    st.plotly_chart(fig, use_container_width=True, key=f"quadrant_{key_suffix}")
    st.caption("Cada burbuja es una IPS. Eje X = % de contraprestación que cobra (menor = más barato para el estudiante), eje Y = score multicriterio (mayor = mejor), tamaño = cupos efectivamente asignados. El cuadrante **arriba-izquierda** agrupa la frontera eficiente: alto score y bajo costo. Las líneas punteadas son las medianas de cada eje.")


def main():
    """Función principal"""
    render_header()

    # Sidebar simplificado
    with st.sidebar:
        st.header("🎯 Opciones")
        st.caption("La página muestra Entrada + Resultados en un solo flujo.")

    # Entrada
    uploaded_file = render_upload_section()

    if uploaded_file is None:
        st.info("ℹ️ Carga primero el Excel para ver Set_ID y Semestres disponibles.")
        return

    # Configuración dinámica según el archivo subido
    set_options, semestre_options = get_config_options_from_upload(uploaded_file)
    default_set = "SET-MEDICINA" if "SET-MEDICINA" in set_options else (set_options[0] if set_options else "SET001")
    default_sem = "2026-1" if "2026-1" in semestre_options else (semestre_options[0] if semestre_options else "2026-1")

    st.markdown("---")
    modo = st.radio(
        "Modo de optimización",
        ["Agregado (actual)", "Refinado por semestre"],
        horizontal=True,
        index=1,
    )

    set_id = default_set
    semestre = default_sem
    total_estudiantes = 80
    programa_manual = "Medicina"
    tipo_est_manual = "Pregrado"
    tipo_practica_manual = "Rotación pregrado"

    if modo == "Agregado (actual)":
        (
            set_id,
            semestre,
            total_estudiantes,
            programa_manual,
            tipo_est_manual,
            tipo_practica_manual,
        ) = render_config_section(
            set_options=set_options,
            semestre_options=semestre_options,
            default_set=default_set,
            default_semestre=default_sem,
        )

    selecciones_refinado = []
    n_por_semestre = {}

    if modo == "Refinado por semestre":
        st.subheader("⚙️ Configuración Refinada (multi-semestre)")
        st.caption(
            "Selecciona una o varias asignaturas de **cualquier semestre**. "
            "A cada asignatura puedes asignarle un set de ponderaciones distinto."
        )

        sem_set_map = {
            5: "SET-SEM5-SaludPublica",
            6: "SET-SEM6-Psiquiatria",
            7: "SET-SEM7-MedicinaInterna",
            8: "SET-SEM8-Pediatria",
            9: "SET-SEM9-Gineco",
            10: "SET-SEM10-Cirugia",
        }

        # Construir catálogo de todas las (semestre, asignatura) disponibles
        opciones_labels = []
        label_to_pair = {}
        for sem in [5, 6, 7, 8, 9, 10]:
            for asig in get_asignaturas_for_semestre(uploaded_file, sem):
                label = f"Sem {sem} · {asig}"
                opciones_labels.append(label)
                label_to_pair[label] = (sem, asig)

        if not opciones_labels:
            st.warning("⚠️ No se encontraron asignaturas en 06_Rotaciones. Verifica el archivo.")
        else:
            seleccion_labels = st.multiselect(
                "Asignaturas a optimizar",
                options=opciones_labels,
                default=[],
                help="Puedes mezclar asignaturas de distintos semestres en una misma corrida.",
                key="seleccion_asignaturas_multi",
            )

            if not seleccion_labels:
                st.info("ℹ️ Selecciona al menos una asignatura para continuar.")
            else:
                st.markdown("##### 🎛️ Set de ponderaciones por asignatura")
                for label in seleccion_labels:
                    sem, asig = label_to_pair[label]
                    c1, c2 = st.columns([3, 2])
                    c1.markdown(f"**Sem {sem}** · {asig}")
                    default_set = sem_set_map.get(sem, "SET-MEDICINA")
                    if default_set not in set_options and set_options:
                        default_set = set_options[0]
                    idx_default = set_options.index(default_set) if default_set in set_options else 0
                    set_sel = c2.selectbox(
                        "Set",
                        options=set_options,
                        index=idx_default,
                        key=f"set_{sem}_{asig}",
                        label_visibility="collapsed",
                    )
                    selecciones_refinado.append(
                        {"semestre": sem, "asignatura": asig, "set_id": set_sel}
                    )

                # Demanda por semestre (default desde 07_Demanda_Semestres)
                st.markdown("##### 👥 Estudiantes por semestre")
                semestres_usados = sorted({s["semestre"] for s in selecciones_refinado})
                cols_dem = st.columns(min(len(semestres_usados), 4) or 1)
                for i, sem in enumerate(semestres_usados):
                    constraints = get_group_constraints(sem)
                    demanda_default = get_demanda_for_semestre(uploaded_file, sem) or 60
                    with cols_dem[i % len(cols_dem)]:
                        n_por_semestre[sem] = st.number_input(
                            f"Sem {sem} (grupos {constraints['min']}-{constraints['max']})",
                            min_value=constraints["min"],
                            value=int(demanda_default),
                            step=1,
                            key=f"n_estudiantes_sem_{sem}",
                        )

    else:
        capacidad_total = preview_capacidad(uploaded_file)
        c1, c2, c3 = st.columns(3)
        c1.metric("Estudiantes a asignar", int(total_estudiantes))
        c2.metric("Cupos disponibles (estimado)", int(capacidad_total))
        c3.metric("Brecha preliminar", int(total_estudiantes - capacidad_total))

        if capacidad_total <= 0:
            st.error("❌ No se detectaron cupos válidos en 02_Oferta_x_Programa.")
        elif int(total_estudiantes) > int(capacidad_total):
            st.warning("⚠️ La cantidad de estudiantes es mayor que la capacidad total disponible; la cobertura será parcial.")
        else:
            st.success("✅ La capacidad total alcanza para la demanda indicada (a nivel agregado).")

    # Bloquear ejecución en modo refinado si no hay asignaturas seleccionadas
    _puede_ejecutar = True
    if modo == "Refinado por semestre" and not selecciones_refinado:
        _puede_ejecutar = False

    if st.button(
        "🚀 Ejecutar Optimización",
        use_container_width=True,
        type="primary",
        disabled=not _puede_ejecutar,
    ):
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(uploaded_file.getbuffer())
            tmp_path = tmp.name

        try:
            with st.spinner("⏳ Procesando..."):
                set_id_to_use = set_id if modo != "Refinado por semestre" else (
                    selecciones_refinado[0]["set_id"] if selecciones_refinado else set_id
                )
                loader = DataLoader(tmp_path, set_id_to_use, semestre)
                loader.load_all()
                st.session_state.loader = loader

                if modo == "Refinado por semestre":
                    if loader.rotaciones is None or loader.rotaciones.empty:
                        st.error("❌ El archivo no contiene la hoja '06_Rotaciones'. Usa Plantilla_V4_Refinada.xlsx")
                        st.session_state.results = None
                    else:
                        st.session_state.results = procesar_refinado(
                            loader,
                            selecciones_refinado,
                            n_por_semestre,
                            semestre,
                        )
                        st.session_state.modo_resultado = "refinado"
                else:
                    st.session_state.results = procesar_datos(
                        loader,
                        set_id,
                        semestre,
                        total_estudiantes,
                        programa_manual,
                        tipo_est_manual,
                        tipo_practica_manual,
                    )
                    st.session_state.modo_resultado = "agregado"

            if st.session_state.results:
                st.success("✅ Optimización completada")
        finally:
            Path(tmp_path).unlink()

    # Resultados en la misma página
    if st.session_state.results:
        st.markdown("---")
        results = st.session_state.results
        modo_res = st.session_state.get("modo_resultado", "agregado")

        if modo_res == "refinado":
            st.header("📊 Resultados — Modelo Refinado (multi-semestre)")

            df_asig = results["asignaciones"]
            por_sem = results.get("por_semestre", {})

            # KPIs globales
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Semestres", len(por_sem))
            col2.metric("Asignaturas", len({(d["asignaturas"]) and a for s, d in por_sem.items() for a in d["asignaturas"]}))
            col3.metric("Estudiantes", results["total_estudiantes"])
            col4.metric("Grupos formados", results["n_grupos_total"])

            # Descarga (Excel completo)
            excel_refinado_bytes = generar_excel_refinado(results)
            st.download_button(
                label="📥 Descargar Excel completo (Asignaciones + Indicadores)",
                data=excel_refinado_bytes,
                file_name="asignaciones_refinado_multisemestre.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_asignaciones_refinado",
            )

            # Indicadores demanda/oferta (vista rápida)
            indic = results.get("indicadores", [])
            if indic:
                st.subheader("📊 Indicadores Demanda / Oferta")
                df_ind = pd.DataFrame(indic)
                df_ind["Ocupación"] = (df_ind["Pct_Demanda_Oferta"] * 100).round(1)
                def _estado(p):
                    if p < 50: return "🟡 Subutilización"
                    if p <= 75: return "🟢 Óptimo"
                    if p <= 100: return "🟡 Alto"
                    return "🔴 Alerta"
                df_ind["Estado"] = df_ind["Ocupación"].apply(_estado)
                st.dataframe(
                    df_ind[["Semestre", "Asignatura", "Set", "Demanda", "Oferta_Maxima", "Ocupación", "Estado"]],
                    use_container_width=True, hide_index=True,
                    column_config={
                        "Ocupación": st.column_config.ProgressColumn(
                            "Ocupación (Dem/Oferta)", format="%.1f%%", min_value=0, max_value=100,
                        ),
                    },
                )

            import plotly.express as px

            # Detalle por semestre (organizado en pestañas)
            if por_sem:
                st.subheader("🗂️ Detalle por semestre")
                tabs = st.tabs([f"Semestre {s}" for s in sorted(por_sem.keys())])
                for tab, sem in zip(tabs, sorted(por_sem.keys())):
                    with tab:
                        d = por_sem[sem]
                        df_s = df_asig[df_asig["Semestre"] == sem].copy()

                        cA, cB, cC, cD = st.columns(4)
                        cA.metric("Estudiantes", d["n_estudiantes"])
                        cB.metric("Asignados", d["asignados"])
                        cC.metric("Grupos", d["n_grupos"])
                        cD.metric("Asignaturas", len(d["asignaturas"]))

                        display_cols = ["Grupo_ID", "Tamano_Grupo", "Asignatura", "Set", "Rotacion",
                                        "ID_Institucion", "Institucion", "Estudiantes", "Score_IPS"]
                        display_cols = [c for c in display_cols if c in df_s.columns]
                        st.dataframe(df_s[display_cols], use_container_width=True, hide_index=True)

                        # Distribución por asignatura del semestre
                        for asig in d["asignaturas"]:
                            df_a = df_s[df_s["Asignatura"] == asig]
                            if df_a.empty:
                                continue
                            label_col = "Institucion" if "Institucion" in df_a.columns else "ID_Institucion"
                            fig = px.bar(
                                df_a, x=label_col, y="Estudiantes", color="Grupo_ID",
                                barmode="stack", title=f"📊 {asig}",
                            )
                            fig.update_xaxes(tickangle=45)
                            st.plotly_chart(fig, use_container_width=True, key=f"dist_s{sem}_{asig}")

            # Análisis avanzado global
            if not df_asig.empty:
                st.subheader("📈 Análisis Avanzado (global)")
                _key = "multi_" + "_".join(str(s) for s in sorted(por_sem.keys()))
                _render_heatmap_grupo_ips(df_asig, key_suffix=_key)
                _render_quadrant_calidad_costo(df_asig, results, st.session_state.loader, key_suffix=_key)
            else:
                st.warning("No se encontraron asignaciones factibles.")
        else:
            st.header("📊 Resultados")
            render_results_summary(results)

            if results["asignaciones"].empty:
                st.error("❌ No hubo asignaciones en esta corrida.")
                st.info(
                    "Revisa el panel de debug: si `pares_factibles` es 0, suele indicar desalineación entre cupos y configuración manual; "
                    "si `pares_con_costo` es muy bajo, faltan costos por institución o tipo de práctica."
                )
            else:
                st.success("✅ Se encontraron asignaciones factibles y se muestran abajo.")

            col1, col2 = st.columns(2)
            with col1:
                render_asignaciones_table(results["asignaciones"])
            with col2:
                render_demanda_vs_asignacion(results["summary"])

            render_capacidad_chart(results["util"])
            render_debug_info(results["debug"])

            st.header("📥 Descargar Resultados")
            excel_bytes = generar_excel_resultados(results)
            st.download_button(
                label="📊 Descargar resultados (Excel)",
                data=excel_bytes,
                file_name="asignaciones_optimizacion.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    # Ayuda siempre disponible abajo
    st.markdown("---")
    st.header("❓ Ayuda")

    with st.expander("🧮 Modelo matemático (MILP)"):
        st.markdown(
            r"""
            ### 1) Tipo de problema
            Este problema se modela como un **MILP de asignación/transportación** con múltiples criterios.

            ### 2) Conjuntos e índices
            - $J$: conjunto de instituciones (IPS/escenarios).
            - $G$: conjunto de grupos de demanda (combinación de programa, tipo de estudiante, tipo de práctica, semestre).
            - $K$: conjunto de criterios activos en el Set seleccionado.

            ### 3) Parámetros
            - $D_g$: estudiantes a asignar en el grupo $g\in G$.
            - $Cap_{j,p,n,s}$: cupo disponible en institución $j$ para programa $p$, nivel $n$, semestre $s$.
            - $w_k$: peso del criterio $k\in K$, con condición $\sum_{k\in K} w_k = 1$.
            - $s_{j,g,k}\in[0,1]$: score normalizado del criterio $k$ para el par $(j,g)$.
            - $V_{j,g}$: score agregado de asignar grupo $g$ a institución $j$.

            ### 4) Construcción del score multicriterio
            Primero se normalizan variables de distinta escala a $[0,1]$:
            - Escala 1–5: $s=(x-1)/4$
            - Escala 0–5: $s=x/5$
            - Porcentaje 0–100: $s=x/100$
            - Binarios: se usan como 0/1

            Para criterios de costo se invierte el sentido (menor costo = mayor score), p. ej.:
            $$
            s^{costo}=1-\frac{x}{100}
            $$

            El score final por par es:
            $$
            V_{j,g}=\sum_{k\in K} w_k\,s_{j,g,k}
            $$

            ### 5) Variable de decisión
            $$
            x_{j,g}\in\mathbb{Z}_{\ge 0}
            $$
            donde $x_{j,g}$ = número de estudiantes del grupo $g$ asignados a la institución $j$.

            ### 6) Función objetivo
            $$
            \max \sum_{g\in G}\sum_{j\in J} V_{j,g}\,x_{j,g}
            $$
            Se maximiza la utilidad total ponderada de la asignación.

            ### 7) Restricciones
            **(a) Satisfacción de demanda por grupo**
            $$
            \sum_{j\in J} x_{j,g} = D_g\qquad \forall g\in G
            $$

            **(b) Capacidad institucional**
            $$
            \sum_{g\in G(p,n,s)} x_{j,g} \le Cap_{j,p,n,s}
            \qquad \forall (j,p,n,s)
            $$

            **(c) Integralidad y no negatividad**
            $$
            x_{j,g}\in\mathbb{Z}_{\ge0}
            $$

            ### 8) Consideraciones prácticas implementadas
            - Si en cupos faltan programa/tipo/semestre, la app aplica fallback con la configuración manual.
            - Si faltan costos para una combinación exacta, se usan reglas de fallback para no perder factibilidad innecesariamente.
            - Si el Set activo no suma 1.0 en pesos, el modelo se bloquea por consistencia metodológica.

            ### 9) Extensiones OR recomendadas (futuras)
            - Límite de fragmentación por grupo: $\sum_j y_{j,g}\le L_g$.
            - Penalización por sobreconcentración en pocas instituciones.
            - Restricciones de equidad (balance entre IPS o entre subgrupos).
            - Frontera eficiencia costo-calidad (análisis multiobjetivo).
            """
        )

    with st.expander("📖 ¿Cómo usar la aplicación?"):
        st.markdown("""
        1. **Cargar archivo**: Sube tu plantilla V3 en Excel
        2. **Configurar**: Selecciona Set_ID, semestre y demanda manual
        3. **Ejecutar**: Haz clic en "Ejecutar Optimización"
        4. **Ver resultados**: Aparecen justo debajo, en esta misma pantalla
        """)

    with st.expander("📋 Estructura de la Plantilla"):
        st.markdown("""
        - **01_Oferta**: Instituciones, servicios y ubicación
        - **02_Oferta_x_Programa**: Cupos disponibles
        - **03_Calidad**: Criterios de calidad
        - **04_Costo_del_Sitio**: Contraprestación y EPP
        - **05_Ponderaciones**: Pesos de criterios
        - **Demanda Pregrado/Posgrado**: Estudiantes a ubicar (opcional)
        """)

    with st.expander("🎯 Criterios Disponibles"):
        criteria = [
            "Es_Hospital_Universitario",
            "Escenario_Avalado_Practicas",
            "Servicios_Pediatricos",
            "Servicios_Obstetricia",
            "MisionVisionProposito_AlineacionDocencia",
            "Admiten_Docentes_Externos",
            "Areas_Bienestar",
            "Areas_Academicas",
            "%_Contraprestacion_Matricula",
            "EPP_Exigidos / Cobro_EPP",
        ]
        for c in criteria:
            st.write(f"• {c}")
if __name__ == "__main__":
    main()
